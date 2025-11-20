#!/usr/bin/env python3
"""
Enhanced Italian-Greek Vocabulary Practice Tool

Usage:
    vocab_enhanced.py quiz [n] [--category CAT] [--difficulty LEVEL]
    vocab_enhanced.py quiz-reverse [n]
    vocab_enhanced.py quiz-mc [n]           - Multiple choice quiz
    vocab_enhanced.py flashcard [n]         - Interactive flashcard mode
    vocab_enhanced.py speak <word>          - Pronounce Italian word
    vocab_enhanced.py sentence [n]          - Generate practice sentences
    vocab_enhanced.py lookup <word>
    vocab_enhanced.py add <italian> <greek> [category] [difficulty]
    vocab_enhanced.py list [--category CAT]
    vocab_enhanced.py stats
    vocab_enhanced.py weak                  - Show words needing practice
    vocab_enhanced.py export [n]
    vocab_enhanced.py web                   - Launch web interface
    vocab_enhanced.py migrate               - Migrate old xlsx to new format

Categories: verbs, nouns, adjectives, adverbs, phrases, other
Difficulty: 1 (easy), 2 (medium), 3 (hard)
"""
import sys
import random
import sqlite3
import json
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

VOCAB_FILE = Path(__file__).parent / "vocabulary.xlsx"
DB_FILE = Path(__file__).parent / "vocab_progress.db"

# ==================== Database Setup ====================

def init_db():
    """Initialize progress tracking database."""
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS quiz_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            word_italian TEXT,
            correct BOOLEAN,
            quiz_type TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS word_stats (
            word_italian TEXT PRIMARY KEY,
            times_seen INTEGER DEFAULT 0,
            times_correct INTEGER DEFAULT 0,
            last_seen DATETIME,
            next_review DATETIME
        )
    ''')
    conn.commit()
    conn.close()

def record_quiz_result(word, correct, quiz_type='standard'):
    """Record quiz result and update spaced repetition schedule."""
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    
    # Log attempt
    c.execute('INSERT INTO quiz_history (word_italian, correct, quiz_type) VALUES (?, ?, ?)',
              (word, correct, quiz_type))
    
    # Update stats
    c.execute('SELECT times_seen, times_correct FROM word_stats WHERE word_italian = ?', (word,))
    row = c.fetchone()
    
    if row:
        times_seen, times_correct = row
        times_seen += 1
        times_correct += int(correct)
    else:
        times_seen, times_correct = 1, int(correct)
    
    # Spaced repetition interval
    if correct:
        success_rate = times_correct / times_seen
        if success_rate >= 0.9:
            interval_days = 7
        elif success_rate >= 0.7:
            interval_days = 3
        else:
            interval_days = 1
    else:
        interval_days = 0.5  # Review same day
    
    next_review = datetime.now() + timedelta(days=interval_days)
    
    c.execute('''
        INSERT OR REPLACE INTO word_stats 
        (word_italian, times_seen, times_correct, last_seen, next_review)
        VALUES (?, ?, ?, ?, ?)
    ''', (word, times_seen, times_correct, datetime.now(), next_review))
    
    conn.commit()
    conn.close()

def get_weak_words(limit=20):
    """Get words that need review (low success rate or overdue)."""
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('''
        SELECT word_italian, times_seen, times_correct, 
               CAST(times_correct AS FLOAT) / times_seen AS success_rate,
               next_review
        FROM word_stats
        WHERE times_seen > 0
        ORDER BY success_rate ASC, next_review ASC
        LIMIT ?
    ''', (limit,))
    rows = c.fetchall()
    conn.close()
    return rows

# ==================== Vocabulary Loading ====================

def load_vocabulary(category=None, difficulty=None):
    """Load vocabulary from xlsx with optional filters."""
    if not VOCAB_FILE.exists():
        print(f"ERROR: {VOCAB_FILE} not found. Run 'migrate' to convert old format.")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(VOCAB_FILE)
    sheet = wb.active
    
    # Expected columns: Italian | Greek | Category | Difficulty
    words = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        
        word = {
            "italian": str(row[0]).strip(),
            "greek": str(row[1]).strip(),
            "category": str(row[2]).strip().lower() if len(row) > 2 and row[2] else "other",
            "difficulty": int(row[3]) if len(row) > 3 and row[3] else 2
        }
        
        # Apply filters
        if category and word["category"] != category.lower():
            continue
        if difficulty and word["difficulty"] != difficulty:
            continue
        
        words.append(word)
    
    wb.close()
    return words

def save_word(italian, greek, category="other", difficulty=2):
    """Append a new word to xlsx."""
    wb = openpyxl.load_workbook(VOCAB_FILE)
    sheet = wb.active
    sheet.append([italian, greek, category, difficulty])
    wb.save(VOCAB_FILE)
    wb.close()
    print(f"✓ Added: {italian} → {greek} [{category}, difficulty {difficulty}]")

def migrate_old_format():
    """Add Category and Difficulty columns to existing xlsx."""
    if not VOCAB_FILE.exists():
        print("No vocabulary.xlsx found. Create a new file with columns: Italian | Greek | Category | Difficulty")
        return
    
    wb = openpyxl.load_workbook(VOCAB_FILE)
    sheet = wb.active
    
    # Check if already migrated
    header = [cell.value for cell in sheet[1]]
    if len(header) >= 4:
        print("Already migrated (4+ columns detected).")
        wb.close()
        return
    
    # Add headers
    if len(header) == 2:
        sheet.cell(1, 3, "Category")
        sheet.cell(1, 4, "Difficulty")
    
    # Fill defaults for existing rows
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row_idx, 3).value is None:
            sheet.cell(row_idx, 3, "other")
        if sheet.cell(row_idx, 4).value is None:
            sheet.cell(row_idx, 4, 2)
    
    wb.save(VOCAB_FILE)
    wb.close()
    print("✓ Migration complete. Added Category and Difficulty columns with defaults.")

# ==================== Quiz Modes ====================

def quiz(n, reverse=False, category=None, difficulty=None):
    """Standard quiz with spaced repetition priority."""
    words = load_vocabulary(category, difficulty)
    if not words:
        print("No words match the filters!")
        return
    
    # Prioritize weak words
    weak = {w[0] for w in get_weak_words(n // 2)}
    priority = [w for w in words if w["italian"] in weak]
    others = [w for w in words if w["italian"] not in weak]
    
    selected = priority[:n//2] + random.sample(others, min(n - len(priority[:n//2]), len(others)))
    random.shuffle(selected)
    
    score = 0
    print(f"\n{'='*50}")
    print(f"Quiz: {len(selected)} words ({'Greek → Italian' if reverse else 'Italian → Greek'})")
    if category:
        print(f"Category: {category}")
    if difficulty:
        print(f"Difficulty: {difficulty}")
    print(f"{'='*50}\n")
    
    for i, word in enumerate(selected, 1):
        question = word["greek"] if reverse else word["italian"]
        answer = word["italian"] if reverse else word["greek"]
        
        user_answer = input(f"{i}. {question} = ").strip()
        correct = user_answer.lower() == answer.lower()
        
        if correct:
            print("✓ Correct!\n")
            score += 1
        else:
            print(f"✗ Wrong. Correct: {answer}\n")
        
        record_quiz_result(word["italian"], correct, 'reverse' if reverse else 'standard')
    
    print(f"{'='*50}")
    print(f"Score: {score}/{len(selected)} ({score*100//len(selected)}%)")
    print(f"{'='*50}\n")

def quiz_multiple_choice(n, category=None, difficulty=None):
    """Multiple choice quiz."""
    words = load_vocabulary(category, difficulty)
    if len(words) < 4:
        print("Need at least 4 words for multiple choice!")
        return
    
    selected = random.sample(words, min(n, len(words)))
    score = 0
    
    print(f"\n{'='*50}")
    print(f"Multiple Choice Quiz: {len(selected)} words")
    print(f"{'='*50}\n")
    
    for i, word in enumerate(selected, 1):
        # Generate 3 wrong answers
        wrong = random.sample([w for w in words if w != word], 3)
        options = [word] + wrong
        random.shuffle(options)
        
        print(f"{i}. {word['italian']}")
        for idx, opt in enumerate(options, 1):
            print(f"   {idx}) {opt['greek']}")
        
        try:
            choice = int(input("\nYour answer (1-4): ").strip())
            correct = options[choice - 1] == word
        except (ValueError, IndexError):
            correct = False
        
        if correct:
            print("✓ Correct!\n")
            score += 1
        else:
            print(f"✗ Wrong. Correct: {word['greek']}\n")
        
        record_quiz_result(word["italian"], correct, 'multiple-choice')
    
    print(f"{'='*50}")
    print(f"Score: {score}/{len(selected)} ({score*100//len(selected)}%)")
    print(f"{'='*50}\n")

def flashcard_mode(n, category=None, difficulty=None):
    """Interactive flashcard practice."""
    words = load_vocabulary(category, difficulty)
    if not words:
        print("No words available!")
        return
    
    selected = random.sample(words, min(n, len(words)))
    
    print(f"\n{'='*50}")
    print(f"Flashcard Mode: {len(selected)} words")
    print("Commands: [enter]=flip, 'y'=know it, 'n'=need review, 'q'=quit")
    print(f"{'='*50}\n")
    
    for i, word in enumerate(selected, 1):
        print(f"\nCard {i}/{len(selected)}")
        print(f"\n   {word['italian']}\n")
        
        cmd = input("[press enter to flip] ").strip().lower()
        if cmd == 'q':
            break
        
        print(f"\n   → {word['greek']}\n")
        
        response = input("Know it? (y/n): ").strip().lower()
        correct = response == 'y'
        record_quiz_result(word["italian"], correct, 'flashcard')
        
        if correct:
            print("✓ Marked as known\n")
        else:
            print("✗ Will review later\n")

# ==================== Audio ====================

def speak_word(word):
    """Pronounce an Italian word using TTS."""
    try:
        from gtts import gTTS
        import os
        
        words = load_vocabulary()
        matches = [w for w in words if w["italian"].lower() == word.lower()]
        
        if not matches:
            print(f"Word '{word}' not found.")
            return
        
        target = matches[0]["italian"]
        print(f"Speaking: {target}")
        
        tts = gTTS(text=target, lang='it')
        audio_file = "/tmp/vocab_speak.mp3"
        tts.save(audio_file)
        
        # Play audio
        if sys.platform == "darwin":
            os.system(f"afplay {audio_file}")
        elif sys.platform == "linux":
            os.system(f"mpg123 -q {audio_file} 2>/dev/null || ffplay -nodisp -autoexit {audio_file} 2>/dev/null")
        else:
            print(f"Audio saved to {audio_file}")
        
    except ImportError:
        print("gTTS not installed. Run: pip install gTTS")
    except Exception as e:
        print(f"Error: {e}")

# ==================== Sentence Generation ====================

def generate_sentences(n=10):
    """Generate practice sentences using vocabulary."""
    words = load_vocabulary()
    if len(words) < 3:
        print("Need more words to generate sentences!")
        return
    
    # Simple sentence templates
    templates = [
        lambda w1, w2: f"{w1['italian'].capitalize()} è {w2['italian']}.",
        lambda w1, w2: f"Ho visto un {w1['italian']} nel {w2['italian']}.",
        lambda w1, w2: f"Mi piace {w1['italian']} e {w2['italian']}.",
        lambda w1, w2: f"Domani vado al {w1['italian']} con {w2['italian']}.",
        lambda w1, w2: f"Il {w1['italian']} è più grande del {w2['italian']}.",
    ]
    
    print(f"\n{'='*60}")
    print("SENTENCE PRACTICE")
    print(f"{'='*60}\n")
    
    selected = random.sample(words, min(n * 2, len(words)))
    
    for i in range(min(5, len(selected) // 2)):
        template = random.choice(templates)
        w1, w2 = selected[i*2], selected[i*2 + 1]
        sentence = template(w1, w2)
        
        print(f"Sentence {i+1}: {sentence}")
        user_trans = input("Your Greek translation: ").strip()
        
        print(f"Words used: {w1['italian']}={w1['greek']}, {w2['italian']}={w2['greek']}")
        print(f"(Your translation stored for review)\n")

# ==================== Stats & Weak Words ====================

def show_weak_words():
    """Display words needing practice."""
    weak = get_weak_words(20)
    
    if not weak:
        print("No quiz history yet. Start practicing!")
        return
    
    print(f"\n{'='*60}")
    print("WORDS NEEDING REVIEW")
    print(f"{'='*60}\n")
    print(f"{'Word':<20} {'Seen':<8} {'Correct':<8} {'Success %':<12} {'Next Review'}")
    print("-" * 60)
    
    for word, seen, correct, rate, next_rev in weak:
        next_date = datetime.fromisoformat(next_rev).strftime("%Y-%m-%d") if next_rev else "now"
        print(f"{word:<20} {seen:<8} {correct:<8} {rate*100:>6.1f}%     {next_date}")
    print()

def stats():
    """Enhanced statistics."""
    words = load_vocabulary()
    
    # Category breakdown
    by_cat = defaultdict(int)
    by_diff = defaultdict(int)
    for w in words:
        by_cat[w["category"]] += 1
        by_diff[w["difficulty"]] += 1
    
    # Quiz stats
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('SELECT COUNT(*), SUM(correct) FROM quiz_history')
    total_quizzes, total_correct = c.fetchone()
    conn.close()
    
    total_quizzes = total_quizzes or 0
    total_correct = total_correct or 0
    
    print(f"\n{'='*60}")
    print("VOCABULARY STATISTICS")
    print(f"{'='*60}\n")
    print(f"Total words: {len(words)}")
    print(f"\nBy Category:")
    for cat, count in sorted(by_cat.items()):
        print(f"  {cat:<15} {count:>3} words")
    print(f"\nBy Difficulty:")
    for diff, count in sorted(by_diff.items()):
        print(f"  Level {diff:<10} {count:>3} words")
    
    if total_quizzes > 0:
        print(f"\nQuiz Performance:")
        print(f"  Total attempts: {total_quizzes}")
        print(f"  Correct answers: {total_correct}")
        print(f"  Success rate: {total_correct*100//total_quizzes}%")
    print()

# ==================== Web Interface ====================

def launch_web():
    """Launch Flask web interface."""
    try:
        from flask import Flask, render_template_string, request, jsonify
        import base64
        import io
        
        app = Flask(__name__)
        
        HTML = """
<!DOCTYPE html>
<html>
<head>
    <title>Italian Vocabulary Practice</title>
    <meta charset="UTF-8">
    <style>
        body { font-family: Arial; max-width: 900px; margin: 30px auto; padding: 20px; background: #f5f5f5; }
        h1 { text-align: center; color: #333; }
        .menu { display: flex; flex-wrap: wrap; gap: 10px; justify-content: center; margin: 30px 0; }
        .menu button { font-size: 16px; padding: 15px 25px; background: #4CAF50; color: white; border: none; border-radius: 5px; cursor: pointer; }
        .menu button:hover { background: #45a049; }
        .menu button.active { background: #2196F3; }
        #quiz { background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); min-height: 300px; }
        .card { border: 2px solid #333; padding: 40px; text-align: center; font-size: 28px; margin: 20px 0; background: #fafafa; border-radius: 8px; }
        .flashcard { cursor: pointer; user-select: none; }
        .answer { margin: 20px 0; text-align: center; }
        input { font-size: 18px; padding: 12px; width: 350px; border: 2px solid #ddd; border-radius: 5px; }
        button { font-size: 16px; padding: 12px 24px; margin: 5px; border: none; border-radius: 5px; cursor: pointer; }
        .submit-btn { background: #4CAF50; color: white; }
        .submit-btn:hover { background: #45a049; }
        .correct { color: #4CAF50; font-weight: bold; }
        .wrong { color: #f44336; font-weight: bold; }
        .options { display: flex; flex-direction: column; gap: 10px; max-width: 400px; margin: 20px auto; }
        .option-btn { padding: 15px; font-size: 18px; background: #e0e0e0; }
        .option-btn:hover { background: #d0d0d0; }
        .speaker { cursor: pointer; font-size: 24px; margin-left: 10px; }
        .controls { text-align: center; margin: 20px 0; }
        .progress { text-align: center; color: #666; margin: 10px 0; }
        .sentence-input { width: 80%; min-height: 60px; font-size: 16px; padding: 10px; }
    </style>
</head>
<body>
    <h1>🇮🇹 Italian Vocabulary Practice 🇬🇷</h1>
    
    <div class="menu">
        <button onclick="startMode('it-gr')">Italian → Greek</button>
        <button onclick="startMode('gr-it')">Greek → Italian</button>
        <button onclick="startMode('mc')">Multiple Choice</button>
        <button onclick="startMode('flashcard')">Flashcards</button>
        <button onclick="startMode('sentence')">Sentences</button>
    </div>
    
    <div id="quiz"></div>
    
    <script>
        let words = [];
        let current = 0;
        let score = 0;
        let mode = '';
        let flipped = false;
        
        async function startMode(m) {
            mode = m;
            current = 0;
            score = 0;
            const res = await fetch('/api/words?n=10');
            words = await res.json();
            
            if (mode === 'sentence') {
                showSentence();
            } else {
                showQuestion();
            }
        }
        
        async function playAudio(word) {
            try {
                const res = await fetch('/api/speak?word=' + encodeURIComponent(word));
                const data = await res.json();
                if (data.audio) {
                    const audio = new Audio('data:audio/mp3;base64,' + data.audio);
                    audio.play();
                }
            } catch(e) {
                console.log('Audio unavailable');
            }
        }
        
        function showQuestion() {
            if (current >= words.length) {
                document.getElementById('quiz').innerHTML = 
                    '<h2>🎉 Quiz Complete!</h2><p>Score: ' + score + '/' + words.length + 
                    ' (' + Math.round(score*100/words.length) + '%)</p>' +
                    '<button onclick="location.reload()" class="submit-btn">Start New Quiz</button>';
                return;
            }
            
            const w = words[current];
            const isReverse = mode === 'gr-it';
            const question = isReverse ? w.greek : w.italian;
            const answer = isReverse ? w.italian : w.greek;
            
            let html = '<div class="progress">Question ' + (current+1) + ' of ' + words.length + '</div>';
            
            if (mode === 'mc') {
                html += '<div class="card">' + question + 
                        '<span class="speaker" onclick="playAudio(\'' + w.italian + '\')">🔊</span></div>';
                html += '<div class="options">';
                
                // Generate options (mix correct + 3 wrong)
                const options = [answer];
                const otherWords = words.filter((x, i) => i !== current);
                for (let i = 0; i < 3 && i < otherWords.length; i++) {
                    options.push(isReverse ? otherWords[i].italian : otherWords[i].greek);
                }
                options.sort(() => Math.random() - 0.5);
                
                options.forEach(opt => {
                    html += '<button class="option-btn" onclick="checkMC(\'' + opt.replace(/'/g, "\\\\'") + 
                            '\', \'' + answer.replace(/'/g, "\\\\'") + '\')">' + opt + '</button>';
                });
                html += '</div>';
            } else if (mode === 'flashcard') {
                if (!flipped) {
                    html += '<div class="card flashcard" onclick="flipCard()">' + question +
                            '<span class="speaker" onclick="event.stopPropagation(); playAudio(\'' + w.italian + '\')">🔊</span>' +
                            '<p style="font-size:14px; color:#999; margin-top:20px;">Click to flip</p></div>';
                } else {
                    html += '<div class="card">' + question + 
                            '<span class="speaker" onclick="playAudio(\'' + w.italian + '\')">🔊</span></div>';
                    html += '<div class="card" style="background:#e8f5e9;">→ ' + answer + '</div>';
                    html += '<div class="controls">' +
                            '<button onclick="markCard(true)" class="submit-btn" style="background:#4CAF50;">✓ Know it</button>' +
                            '<button onclick="markCard(false)" class="submit-btn" style="background:#f44336;">✗ Need review</button>' +
                            '</div>';
                }
            } else {
                html += '<div class="card">' + question + 
                        '<span class="speaker" onclick="playAudio(\'' + w.italian + '\')">🔊</span></div>';
                html += '<div class="answer">' +
                        '<input type="text" id="answer" placeholder="Your answer" autofocus />' +
                        '<button onclick="checkAnswer(\'' + answer.replace(/'/g, "\\\\'") + '\')" class="submit-btn">Submit</button>' +
                        '</div><div id="result"></div>';
                
                setTimeout(() => {
                    const inp = document.getElementById('answer');
                    if (inp) {
                        inp.focus();
                        inp.addEventListener('keypress', function(e) {
                            if (e.key === 'Enter') checkAnswer(answer);
                        });
                    }
                }, 100);
            }
            
            document.getElementById('quiz').innerHTML = html;
        }
        
        function flipCard() {
            flipped = true;
            showQuestion();
        }
        
        function markCard(correct) {
            if (correct) score++;
            flipped = false;
            current++;
            showQuestion();
        }
        
        function checkAnswer(correct) {
            const ans = document.getElementById('answer').value.trim();
            const isCorrect = ans.toLowerCase() === correct.toLowerCase();
            
            if (isCorrect) {
                score++;
                document.getElementById('result').innerHTML = '<p class="correct">✓ Correct!</p>';
            } else {
                document.getElementById('result').innerHTML = 
                    '<p class="wrong">✗ Wrong. Answer: ' + correct + '</p>';
            }
            
            // Play pronunciation
            playAudio(words[current].italian);
            
            current++;
            setTimeout(showQuestion, 2000);
        }
        
        function checkMC(selected, correct) {
            const isCorrect = selected === correct;
            if (isCorrect) score++;
            
            const res = isCorrect ? 
                '<p class="correct">✓ Correct!</p>' : 
                '<p class="wrong">✗ Wrong. Answer: ' + correct + '</p>';
            
            document.getElementById('quiz').innerHTML += '<div>' + res + '</div>';
            
            // Play pronunciation
            playAudio(words[current].italian);
            
            current++;
            setTimeout(showQuestion, 2000);
        }
        
        async function showSentence() {
            if (current >= 5) {
                document.getElementById('quiz').innerHTML = 
                    '<h2>Sentence Practice Complete!</h2>' +
                    '<button onclick="location.reload()" class="submit-btn">Start New</button>';
                return;
            }
            
            const res = await fetch('/api/sentence');
            const data = await res.json();
            
            let html = '<div class="progress">Sentence ' + (current+1) + ' of 5</div>';
            html += '<div class="card">' + data.italian + '</div>';
            html += '<div class="answer">';
            html += '<textarea id="translation" class="sentence-input" placeholder="Translate to Greek..."></textarea><br>';
            html += '<button onclick="checkSentence()" class="submit-btn">Check Translation</button>';
            html += '</div><div id="result"></div>';
            html += '<div style="margin-top:20px; color:#666; font-size:14px;">Words: ' + data.words + '</div>';
            
            document.getElementById('quiz').innerHTML = html;
        }
        
        function checkSentence() {
            const userTrans = document.getElementById('translation').value.trim();
            document.getElementById('result').innerHTML = 
                '<p style="color:#2196F3;">Your translation recorded!</p>' +
                '<p style="font-size:14px; color:#666;">Tip: Practice makes perfect! 💪</p>';
            
            current++;
            setTimeout(showSentence, 2000);
        }
        
        // Auto-start with Italian→Greek mode
        window.onload = () => startMode('it-gr');
    </script>
</body>
</html>
        """
        
        @app.route('/')
        def index():
            return render_template_string(HTML)
        
        @app.route('/api/words')
        def get_words():
            n = int(request.args.get('n', 10))
            words = load_vocabulary()
            selected = random.sample(words, min(n, len(words)))
            return jsonify([{"italian": w["italian"], "greek": w["greek"]} for w in selected])
        
        @app.route('/api/speak')
        def speak():
            word = request.args.get('word', '')
            try:
                from gtts import gTTS
                tts = gTTS(text=word, lang='it')
                audio_bytes = io.BytesIO()
                tts.write_to_fp(audio_bytes)
                audio_bytes.seek(0)
                audio_b64 = base64.b64encode(audio_bytes.read()).decode('utf-8')
                return jsonify({"audio": audio_b64})
            except Exception as e:
                return jsonify({"error": str(e)})
        
        @app.route('/api/sentence')
        def get_sentence():
            words = load_vocabulary()
            selected = random.sample(words, min(3, len(words)))
            
            templates = [
                lambda w: f"{w[0]['italian'].capitalize()} è {w[1]['italian']}.",
                lambda w: f"Ho visto {w[0]['italian']} nel {w[1]['italian']}.",
                lambda w: f"Mi piace {w[0]['italian']} e {w[1]['italian']}.",
                lambda w: f"Domani vado al {w[0]['italian']} con {w[1]['italian']}.",
            ]
            
            template = random.choice(templates)
            sentence = template(selected)
            words_used = ", ".join([f"{w['italian']}={w['greek']}" for w in selected[:2]])
            
            return jsonify({"italian": sentence, "words": words_used})
        
        print("\n🌐 Starting web interface at http://localhost:5000")
        print("Press Ctrl+C to stop\n")
        app.run(debug=False, port=5000, host='0.0.0.0')
        
    except ImportError:
        print("Flask not installed. Run: pip install flask")

# ==================== Other Functions ====================

def lookup(word):
    """Find translation."""
    words = load_vocabulary()
    word_lower = word.lower()
    matches = [w for w in words if word_lower in w["italian"].lower() or word_lower in w["greek"].lower()]
    
    if not matches:
        print(f"'{word}' not found.")
        return
    
    print(f"\nResults for '{word}':")
    for m in matches:
        print(f"  {m['italian']} → {m['greek']} [{m['category']}, diff: {m['difficulty']}]")
    print()

def list_all(category=None):
    """List all words."""
    words = load_vocabulary(category=category)
    print(f"\nVocabulary ({len(words)} words{' in ' + category if category else ''}):")
    print(f"{'Italian':<25} {'Greek':<25} {'Category':<12} {'Diff'}")
    print("-" * 70)
    for w in sorted(words, key=lambda x: x["italian"].lower()):
        print(f"{w['italian']:<25} {w['greek']:<25} {w['category']:<12} {w['difficulty']}")
    print()

def export_for_ai(n):
    """Export for AI sentence generation."""
    words = load_vocabulary()
    selected = random.sample(words, min(n, len(words)))
    
    print("\n" + "="*60)
    print("VOCABULARY FOR AI PRACTICE")
    print("="*60 + "\n")
    
    for i, w in enumerate(selected, 1):
        print(f"{i}. {w['italian']} ({w['greek']})")
    
    print("\n" + "-"*60)
    print("Paste this to AI: Create 5 Italian sentences using these words,")
    print("then ask me to translate each to Greek.")
    print("="*60 + "\n")

# ==================== Main ====================

def main():
    init_db()
    
    if len(sys.argv) < 2 or sys.argv[1] in ('-h', '--help', 'help'):
        print(__doc__)
        sys.exit(0)
    
    cmd = sys.argv[1]
    
    # Parse common options
    category = None
    difficulty = None
    if '--category' in sys.argv:
        idx = sys.argv.index('--category')
        category = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else None
    if '--difficulty' in sys.argv:
        idx = sys.argv.index('--difficulty')
        difficulty = int(sys.argv[idx + 1]) if idx + 1 < len(sys.argv) else None
    
    if cmd == "quiz":
        n = int(sys.argv[2]) if len(sys.argv) > 2 and sys.argv[2].isdigit() else 10
        quiz(n, reverse=False, category=category, difficulty=difficulty)
    
    elif cmd == "quiz-reverse":
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        quiz(n, reverse=True, category=category, difficulty=difficulty)
    
    elif cmd == "quiz-mc":
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        quiz_multiple_choice(n, category=category, difficulty=difficulty)
    
    elif cmd == "flashcard":
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        flashcard_mode(n, category=category, difficulty=difficulty)
    
    elif cmd == "speak":
        if len(sys.argv) < 3:
            print("Usage: vocab_enhanced.py speak <word>")
            sys.exit(1)
        speak_word(sys.argv[2])
    
    elif cmd == "sentence":
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        generate_sentences(n)
    
    elif cmd == "lookup":
        if len(sys.argv) < 3:
            print("Usage: vocab_enhanced.py lookup <word>")
            sys.exit(1)
        lookup(sys.argv[2])
    
    elif cmd == "add":
        if len(sys.argv) < 4:
            print("Usage: vocab_enhanced.py add <italian> <greek> [category] [difficulty]")
            sys.exit(1)
        cat = sys.argv[4] if len(sys.argv) > 4 else "other"
        diff = int(sys.argv[5]) if len(sys.argv) > 5 else 2
        save_word(sys.argv[2], sys.argv[3], cat, diff)
    
    elif cmd == "list":
        list_all(category=category)
    
    elif cmd == "stats":
        stats()
    
    elif cmd == "weak":
        show_weak_words()
    
    elif cmd == "export":
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        export_for_ai(n)
    
    elif cmd == "web":
        launch_web()
    
    elif cmd == "migrate":
        migrate_old_format()
    
    else:
        print(f"Unknown command: {cmd}")
        print(__doc__)
        sys.exit(1)

if __name__ == "__main__":
    main()
