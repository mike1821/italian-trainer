#!/usr/bin/env python3
"""
Core vocabulary management functions.
"""
import openpyxl
import random
from pathlib import Path

VOCAB_FILE = Path("vocabulary.xlsx")

def load_vocabulary():
    """Load vocabulary from Excel file."""
    if not VOCAB_FILE.exists():
        raise FileNotFoundError(f"Vocabulary file not found: {VOCAB_FILE}")
    
    wb = openpyxl.load_workbook(VOCAB_FILE)
    ws = wb.active
    
    words = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        word = {
            "italian": str(row[0]).strip(),
            "greek": str(row[1]).strip() if row[1] else "",
            "category": str(row[2]).strip() if len(row) > 2 and row[2] else "other",
            "difficulty": int(row[3]) if len(row) > 3 and row[3] else 2
        }
        words.append(word)
    
    wb.close()
    
    if not words:
        raise ValueError("No vocabulary words found in the file")
    
    return words


def filter_words(words, category=None, difficulty=None):
    """Filter words by category and/or difficulty."""
    filtered = words
    
    if category:
        filtered = [w for w in filtered if w["category"].lower() == category.lower()]
    
    if difficulty is not None:
        filtered = [w for w in filtered if w["difficulty"] == difficulty]
    
    return filtered


def get_random_words(words, n=10):
    """Get n random words from the vocabulary."""
    return random.sample(words, min(n, len(words)))


def migrate_vocabulary():
    """Migrate old 2-column format to new 4-column format."""
    if not VOCAB_FILE.exists():
        print(f"Error: {VOCAB_FILE} not found")
        return False
    
    wb = openpyxl.load_workbook(VOCAB_FILE)
    ws = wb.active
    
    # Check if already migrated
    if ws.cell(1, 3).value:
        print("Vocabulary already has Category/Difficulty columns")
        wb.close()
        return False
    
    # Add headers
    ws.cell(1, 3, "Category")
    ws.cell(1, 4, "Difficulty")
    
    # Add default values
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value:  # If Italian word exists
            ws.cell(row, 3, "other")  # Default category
            ws.cell(row, 4, 2)        # Default difficulty
    
    wb.save(VOCAB_FILE)
    wb.close()
    
    print(f"✓ Migration complete! Added Category and Difficulty columns.")
    print(f"  Default values: Category='other', Difficulty=2")
    return True
