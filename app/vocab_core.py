#!/usr/bin/env python3
"""
Core vocabulary management functions.
"""
import openpyxl
import random
from pathlib import Path

VOCAB_FILE = Path("vocabulary.xlsx")

def load_vocabulary():
    """Load vocabulary from Excel file."""
    if not VOCAB_FILE.exists():
        raise FileNotFoundError(f"Vocabulary file not found: {VOCAB_FILE}")
    
    wb = openpyxl.load_workbook(VOCAB_FILE)
    ws = wb.active
    
    words = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        word = {
            "italian": str(row[0]).strip(),
            "greek": str(row[1]).strip() if row[1] else "",
            "category": str(row[2]).strip() if len(row) > 2 and row[2] else "other",
            "difficulty": int(row[3]) if len(row) > 3 and row[3] else 2
        }
        words.append(word)
    
    wb.close()
    
    if not words:
        raise ValueError("No vocabulary words found in the file")
    
    return words


def filter_words(words, category=None, difficulty=None):
    """Filter words by category and/or difficulty."""
    filtered = words
    
    if category:
        filtered = [w for w in filtered if w["category"].lower() == category.lower()]
    
    if difficulty is not None:
        filtered = [w for w in filtered if w["difficulty"] == difficulty]
    
    return filtered


def get_random_words(words, n=10):
    """Get n random words from the vocabulary."""
    return random.sample(words, min(n, len(words)))


def migrate_vocabulary():
    """Migrate old 2-column format to new 4-column format, and fill defaults for any rows missing values."""
    if not VOCAB_FILE.exists():
        print(f"Error: {VOCAB_FILE} not found")
        return False
    
    wb = openpyxl.load_workbook(VOCAB_FILE)
    ws = wb.active
    
    added_headers = False
    if not ws.cell(1, 3).value:
        ws.cell(1, 3, "Category")
        ws.cell(1, 4, "Difficulty")
        added_headers = True
    
    filled = 0
    for row in range(2, ws.max_row + 1):
        if not ws.cell(row, 1).value:
            continue
        if not ws.cell(row, 3).value:
            ws.cell(row, 3, "other")
            filled += 1
        if not ws.cell(row, 4).value:
            ws.cell(row, 4, 2)
    
    if not added_headers and filled == 0:
        print("Nothing to migrate — all rows already have Category and Difficulty.")
        wb.close()
        return False
    
    wb.save(VOCAB_FILE)
    wb.close()
    
    if added_headers:
        print(f"✓ Migration complete! Added Category and Difficulty columns.")
    print(f"✓ Filled defaults for {filled} row(s) (Category='other', Difficulty=2).")
    return True
