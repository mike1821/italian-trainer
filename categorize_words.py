#!/usr/bin/env python3
"""
Categorize Italian vocabulary in vocabulary.xlsx using consistent categories:
Nouns, Adjectives, Prepositions, Objects, Phrases, Colors, Numbers, Days of the Week,
Verbs, Adverbs, Conjunctions, Pronouns, Time, Greetings, Family, Food, Weather, Body, Other.
"""
import openpyxl
import re
from pathlib import Path

VOCAB_FILE = Path("vocabulary.xlsx")

# --- Explicit word lists (lowercase for matching) ---

DAYS_OF_THE_WEEK = {
    "lunedì", "martedì", "mercoledì", "giovedì", "venerdì", "sabato", "domenica",
    "lunedi", "martedi", "mercoledi", "giovedi", "venerdi",  # no accent
}

NUMBERS = {
    "zero", "uno", "due", "tre", "quattro", "cinque", "sei", "sette", "otto", "nove", "dieci",
    "undici", "dodici", "tredici", "quattordici", "quindici", "sedici", "diciassette", "diciotto", "diciannove",
    "venti", "trenta", "quaranta", "cinquanta", "sessanta", "settanta", "ottanta", "novanta", "cento",
    "quarnta",  # typo in source
}

COLORS = {
    "bianco", "nero", "rosso", "blu", "giallo", "verde", "marrone", "arancione", "grigio", "rosa", "viola", "azzurro", "beige",
}

PREPOSITIONS = {
    "a", "di", "da", "in", "su", "per", "tra", "fra", "con", "senza", "verso", "sul", "sulla", "sui", "sulle", "nel", "nella", "nello", "negli", "nelle",
    "dopo", "prima", "sopra", "sotto", "dentro", "fuori", "vicino", "lontano", "insieme", "contro", "durante", "oltre", "secondo", "circa",
    "propria", "suo", "mia", "tua",  # possessive / pronominal
}

CONJUNCTIONS = {
    "e", "ed", "ma", "o", "oppure", "né", "perché", "che", "se", "quando", "mentre", "come", "dunque", "quindi", "però", "anzi", "cioè", "infatti", "anche se", "comunque", "allora", "perciò", "qualora", "benchè", "sebbene", "cioe",
}

PRONOUNS = {
    "io", "tu", "lui", "lei", "noi", "voi", "loro", "mi", "ti", "si", "ci", "vi", "lo", "la", "li", "le", "ne",
    "questo", "questa", "quello", "quella", "quale", "quanto", "quanta", "chi", "cui", "niente", "qualcosa", "qualcuno", "ogni",
}

GREETINGS = {
    "ciao", "buongiorno", "buonasera", "buonanotte", "grazie", "prego", "salve", "arrivederci", "a presto", "benvenuto", "benvenuta", "addio", "scusa", "permesso",
}

TIME = {
    "oggi", "domani", "ieri", "ora", "adesso", "presto", "tardi", "poi", "sempre", "mai", "spesso", "subito", "ancora", "già", "finalmente",
    "notte", "mattina", "sera", "pomeriggio", "mezzogiorno", "mezzanotte", "orario", "anticipo", "ritardo",
}

FAMILY = {
    "madre", "padre", "mamma", "papà", "papa", "figlio", "figlia", "fratello", "fratelli", "sorella", "nonno", "nonna", "zio", "zia", "cugino", "cugina", "marito", "moglie",
    "famiglia", "genitori", "bambino", "bambina", "ragazzo", "ragazza", "figlio unico", "fidanzato", "fidanzata",
}

FOOD = {
    "pane", "pasta", "carne", "pesce", "formaggio", "vino", "acqua", "caffè", "caffe", "tè", "te", "latte", "frutta", "frutti", "verdura", "dolce", "zucchero", "sale", "olio", "riso", "pizza", "gelato", "biscotti",
    "cena", "pranzo", "colazione", "arance", "noccioline", "bottiglia",
}

WEATHER = {
    "sole", "pioggia", "neve", "vento", "caldo", "freddo", "nuvola", "cielo", "temperatura", "estate", "inverno", "primavera", "autunno",
}

BODY = {
    "testa", "occhio", "occhi", "orecchio", "naso", "bocca", "mano", "piede", "braccio", "gamba", "cuore", "corpo", "dito", "capelli", "labro",
}

# Physical objects (concrete things) - subset of nouns
OBJECTS = {
    "penna", "chiave", "bicchiere", "tavolo", "sedia", "letto", "porta", "finestra", "libro", "quaderno", "telefono", "orologio",
    "bottiglia", "piatto", "tazza", "cucchiaio", "forchetta", "coltello", "lavagna", "banco", "biglietti", "regalo", "immagine",
    "costellazione", "stella", "pianeti", "aquila", "cane", "gatto", "albero", "giardino", "isola", "paese", "origine",
    "barzellette", "occhiali", "scarpe", "stivale", "tesoro",
}

# Verbs: infinitives and common verb forms (we also use -are/-ere/-ire below)
VERB_EXTRA = {
    "andiamo", "capisco", "esco", "devo", "posso", "sai", "sono", "sto", "ho", "hai", "dai", "scendi", "supera", "portali",
    "apri", "entra", "apre", "aspetto", "ricordi", "ruba",
}


def normalize(s):
    """Normalize for matching: lowercase, strip, collapse space."""
    if not s:
        return ""
    return " ".join(str(s).lower().strip().split())


def is_multi_word(italian):
    return " " in italian.strip() or "/" in italian or "?" in italian or "!" in italian


def is_verb_infinitive(italian):
    """Check if word looks like an Italian infinitive."""
    w = normalize(italian).replace(" ", "")
    if "/" in italian:
        parts = italian.split("/")
        w = normalize(parts[0]).replace(" ", "")
    return w.endswith("are") or w.endswith("ere") or w.endswith("ire")


def is_adverb(italian):
    w = normalize(italian)
    return w.endswith("mente") or w in {
        "molto", "poco", "troppo", "bene", "male", "qui", "qua", "lì", "là", "così", "davvero", "forse", "magari",
        "naturalmente", "finalmente", "quasi", "solo", "anche", "pure", "inoltre", "oltretutto", "certamente",
        "insieme", "abbastanza", "perfino", "almeno", "soltanto", "specialmente", "specialmente", "esattamente",
        "probabilmente", "realmente", "veramente", "sicuramente",
    }


def categorize_word(italian, _greek=""):
    italian_norm = normalize(italian)
    italian_single = italian_norm.split("/")[0].strip().split()[0] if italian_norm else ""

    # 1) Multi-word (or contains / ? !) -> Phrases
    if is_multi_word(italian):
        return "Phrases"

    # 2) Days of the Week
    if italian_single in DAYS_OF_THE_WEEK or italian_norm in DAYS_OF_THE_WEEK:
        return "Days of the Week"

    # 3) Numbers
    if italian_single in NUMBERS or italian_norm in NUMBERS:
        return "Numbers"

    # 4) Colors
    if italian_single in COLORS or italian_norm in COLORS:
        return "Colors"

    # 5) Prepositions
    if italian_single in PREPOSITIONS or italian_norm in PREPOSITIONS:
        return "Prepositions"

    # 6) Conjunctions
    if italian_single in CONJUNCTIONS or italian_norm in CONJUNCTIONS:
        return "Conjunctions"

    # 7) Pronouns
    if italian_single in PRONOUNS or italian_norm in PRONOUNS:
        return "Pronouns"

    # 8) Greetings
    if italian_single in GREETINGS or italian_norm in GREETINGS:
        return "Greetings"

    # 9) Time
    if italian_single in TIME or italian_norm in TIME:
        return "Time"

    # 10) Family
    if italian_single in FAMILY or any(f in italian_norm for f in FAMILY):
        return "Family"

    # 11) Food
    if italian_single in FOOD or italian_norm in FOOD:
        return "Food"

    # 12) Weather
    if italian_single in WEATHER or italian_norm in WEATHER:
        return "Weather"

    # 13) Body
    if italian_single in BODY or italian_norm in BODY:
        return "Body"

    # 14) Verbs (infinitives and extra list)
    if italian_single in VERB_EXTRA or italian_norm in VERB_EXTRA:
        return "Verbs"
    if is_verb_infinitive(italian):
        return "Verbs"

    # 15) Adverbs
    if is_adverb(italian):
        return "Adverbs"

    # 16) Objects (concrete things)
    if italian_single in OBJECTS or italian_norm in OBJECTS:
        return "Objects"

    # 17) Adjectives: common endings or known list
    adj_endings = ("ivo", "oso", "ibile", "abile", "ante", "ente", "ico", "ivo", "ale", "are")
    known_adj = {"alto", "basso", "bello", "brutto", "buono", "cattivo", "nuovo", "vecchio", "giovane", "grande", "piccolo", "lungo", "corto", "facile", "difficile", "importante", "buono", "carino", "contenta", "comodo", "dolce", "diverso", "differente", "caro", "care", "coraggioso", "scortese", "stressato", "entusiasta", "mite", "miglior", "straniera", "piccola", "piccolo", "adatta", "diffuso", "sicuro", "torto", "diverente", "piccola"}
    if italian_single in known_adj or italian_norm in known_adj:
        return "Adjectives"
    if len(italian_single) > 3 and (italian_single.endswith(("o", "a", "e")) or any(italian_single.endswith(e) for e in adj_endings)):
        # Could be adjective or noun; prefer adjective if -issimo or typical adj
        if "issimo" in italian_single or italian_single.endswith(("ivo", "oso", "ibile", "abile")):
            return "Adjectives"
        # Single -o/-a/-e might be noun; we'll leave to Nouns below if not in objects
        pass

    # 18) Nouns (default for remaining single words that look like things/concepts)
    noun_endings = ("zione", "tà", "ità", "anza", "enza", "ore", "trice", "aggio", "mento")
    if any(italian_single.endswith(e) for e in noun_endings):
        return "Nouns"

    # Default: Nouns for single-word items that don't fit elsewhere
    return "Nouns"


def apply_categories_to_excel(backup=True):
    if not VOCAB_FILE.exists():
        print(f"Error: {VOCAB_FILE} not found")
        return

    wb = openpyxl.load_workbook(VOCAB_FILE)
    ws = wb.active

    # Ensure header row has "Category" (column C)
    if ws.max_column < 3:
        ws.cell(1, 3, "Category")
    if ws.cell(1, 3).value is None or str(ws.cell(1, 3).value).strip() == "":
        ws.cell(1, 3, "Category")
    if ws.max_column < 4:
        ws.cell(1, 4, "Difficulty level")

    updates = []
    for row in range(2, ws.max_row + 1):
        italian = ws.cell(row, 1).value
        greek = ws.cell(row, 2).value
        if not italian:
            continue
        italian = str(italian).strip()
        greek = str(greek).strip() if greek else ""
        new_cat = categorize_word(italian, greek)
        old_cat = ws.cell(row, 3).value
        old_cat = str(old_cat).strip() if old_cat else ""
        if new_cat != old_cat:
            updates.append((row, italian, old_cat, new_cat))
        ws.cell(row, 3, new_cat)
        if ws.max_column >= 4 and (ws.cell(row, 4).value is None or ws.cell(row, 4).value == ""):
            try:
                d = int(ws.cell(row, 4).value) if ws.cell(row, 4).value else 2
            except Exception:
                d = 2
            ws.cell(row, 4, d)

    wb.save(VOCAB_FILE)
    wb.close()

    from collections import Counter
    words = []
    wb2 = openpyxl.load_workbook(VOCAB_FILE)
    for row in wb2.active.iter_rows(min_row=2, values_only=True):
        if row[0]:
            words.append(str(row[2]) if len(row) > 2 and row[2] else "Other")
    wb2.close()
    counts = Counter(words)

    print("Categories applied to", VOCAB_FILE)
    print("\nCategory counts:")
    for cat, count in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")
    print(f"\nTotal rows updated: {len(updates)}")
    if updates and len(updates) <= 30:
        for row, it, old, new in updates[:30]:
            print(f"  Row {row}: {it!r}  {old!r} -> {new!r}")
    elif updates:
        print(f"  (First 15 changes: )")
        for row, it, old, new in updates[:15]:
            print(f"  Row {row}: {it!r}  {old!r} -> {new!r}")


if __name__ == "__main__":
    apply_categories_to_excel()
