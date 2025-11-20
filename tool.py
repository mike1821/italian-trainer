#!/usr/bin/env python3
"""
Italian-Greek Vocabulary Practice Tool
Usage:
    python tool.py quiz <n>       - Random quiz with n words (Italian->Greek)
    python tool.py quiz-reverse <n> - Quiz Greek->Italian
    python tool.py lookup <word>  - Find translation
    python tool.py add <italian> <greek> - Add new word to xlsx
    python tool.py list           - Show all words
    python tool.py stats          - Show vocabulary statistics
    python tool.py export <n>     - Export n random words for AI sentence practice
"""
import sys
import random
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

VOCAB_FILE = Path(__file__).parent / "vocabulary.xlsx"


def load_vocabulary():
    """Load vocabulary from xlsx file."""
    if not VOCAB_FILE.exists():
        print(f"ERROR: {VOCAB_FILE} not found")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(VOCAB_FILE)
    sheet = wb.active
    
    # Assuming first row is header: Italian | Greek
    words = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # Skip empty rows
            words.append({"italian": str(row[0]).strip(), "greek": str(row[1]).strip()})
    
    wb.close()
    return words


def save_word(italian, greek):
    """Append a new word to the xlsx file."""
    wb = openpyxl.load_workbook(VOCAB_FILE)
    sheet = wb.active
    sheet.append([italian, greek])
    wb.save(VOCAB_FILE)
    wb.close()
    print(f"✓ Added: {italian} → {greek}")


def quiz(n, reverse=False):
    """Run a quiz with n random words."""
    words = load_vocabulary()
    if not words:
        print("No words in vocabulary!")
        return
    
    n = min(n, len(words))
    selected = random.sample(words, n)
    
    score = 0
    print(f"\n{'='*50}")
    print(f"Quiz: {n} words ({'Greek → Italian' if reverse else 'Italian → Greek'})")
    print(f"{'='*50}\n")
    
    for i, word in enumerate(selected, 1):
        question = word["greek"] if reverse else word["italian"]
        answer = word["italian"] if reverse else word["greek"]
        
        user_answer = input(f"{i}. {question} = ").strip()
        
        if user_answer.lower() == answer.lower():
            print("✓ Correct!\n")
            score += 1
        else:
            print(f"✗ Wrong. Correct answer: {answer}\n")
    
    print(f"{'='*50}")
    print(f"Score: {score}/{n} ({score*100//n}%)")
    print(f"{'='*50}\n")


def lookup(word):
    """Find translation for a word."""
    words = load_vocabulary()
    word_lower = word.lower()
    
    matches = [w for w in words if word_lower in w["italian"].lower() or word_lower in w["greek"].lower()]
    
    if not matches:
        print(f"'{word}' not found in vocabulary.")
        return
    
    print(f"\nResults for '{word}':")
    for m in matches:
        print(f"  {m['italian']} → {m['greek']}")
    print()


def list_all():
    """List all vocabulary words."""
    words = load_vocabulary()
    print(f"\nVocabulary ({len(words)} words):")
    print(f"{'Italian':<30} {'Greek':<30}")
    print("-" * 60)
    for w in sorted(words, key=lambda x: x["italian"].lower()):
        print(f"{w['italian']:<30} {w['greek']:<30}")
    print()


def stats():
    """Show vocabulary statistics."""
    words = load_vocabulary()
    print(f"\nVocabulary Statistics:")
    print(f"  Total words: {len(words)}")
    print(f"  Average Italian word length: {sum(len(w['italian']) for w in words) / len(words):.1f} chars")
    print(f"  Average Greek word length: {sum(len(w['greek']) for w in words) / len(words):.1f} chars")
    print()


def export_for_ai(n):
    """Export n random words in a format optimized for AI sentence generation."""
    words = load_vocabulary()
    if not words:
        print("No words in vocabulary!")
        return
    
    n = min(n, len(words))
    selected = random.sample(words, n)
    
    print("\n" + "="*60)
    print("VOCABULARY FOR AI PRACTICE")
    print("="*60)
    print("\nSelected Italian words with Greek translations:\n")
    
    for i, word in enumerate(selected, 1):
        print(f"{i}. {word['italian']} ({word['greek']})")
    
    print("\n" + "-"*60)
    print("PROMPT FOR AI:")
    print("-"*60)
    print("""
Create 5 Italian sentences using the words above. Requirements:
- Use 2-3 of these words per sentence
- Mix difficulty: simple present, past, future, or conditional tenses
- Include context (events, daily life, descriptions)
- After each sentence, ask me to translate it to Greek
- Then reveal the correct translation and explain any grammar points

Start with: "Sentence 1:" followed by the Italian sentence.
""")
    print("="*60 + "\n")


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ('-h', '--help', 'help'):
        print(__doc__)
        sys.exit(0)
    
    cmd = sys.argv[1]
    
    if cmd == "quiz":
        # Usage: tool.py quiz [n]
        # Run a quiz with n random words (Italian→Greek). Default: 10 words.
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        quiz(n, reverse=False)
    
    elif cmd == "quiz-reverse":
        # Usage: tool.py quiz-reverse [n]
        # Run a reverse quiz (Greek→Italian). Default: 10 words.
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        quiz(n, reverse=True)
    
    elif cmd == "lookup":
        # Usage: tool.py lookup <word>
        # Search for a word in the vocabulary (partial matching).
        if len(sys.argv) < 3:
            print("Usage: tool.py lookup <word>")
            sys.exit(1)
        lookup(sys.argv[2])
    
    elif cmd == "add":
        # Usage: tool.py add <italian> <greek>
        # Add a new word pair to the vocabulary xlsx file.
        if len(sys.argv) < 4:
            print("Usage: tool.py add <italian> <greek>")
            sys.exit(1)
        save_word(sys.argv[2], sys.argv[3])
    
    elif cmd == "list":
        # Usage: tool.py list
        # Display all vocabulary words in alphabetical order.
        list_all()
    
    elif cmd == "stats":
        # Usage: tool.py stats
        # Show vocabulary statistics (total words, average lengths).
        stats()
    
    elif cmd == "export":
        # Usage: tool.py export [n]
        # Export n random words with prompt for AI sentence generation. Default: 10 words.
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        export_for_ai(n)
    
    else:
        print(f"Unknown command: {cmd}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
